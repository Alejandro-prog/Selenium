from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time



driver = webdriver.Edge(executable_path="C:/cromiun_driver/msedgedriver.exe")


filesheet = "./people.xlsx"
wb = load_workbook(filesheet)
hojas = wb.get_sheet_names() 
print(hojas)
nombres = wb.get_sheet_by_name('Hoja1')
wb.close()

for x in range(1,5):
    driver.get("https://docs.google.com/forms/d/e/1FAIpQLSfaYXnV0o9nCIuF3juVZmi349TM7mWGz_KnoWgLKwn319Kw4Q/viewform?usp=sf_link")
    nombre = nombres.cell(row=x, column=1)
    grupo = nombres.cell(row=x, column=3)
    print(nombre.value, grupo.value)

    time.sleep(1)

    nombre1 = driver.find_element_by_xpath("//*[@id='mG61Hd']/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div/div[1]/input")
    nombre1.send_keys(nombre.value)

    time.sleep(1)

    asistir1 = driver.find_element_by_xpath("//*[@id='i9']/div[3]/div")
    asistir1.click()

    time.sleep(1)

    grupo1 = driver.find_element_by_xpath("//*[@id='mG61Hd']/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div/div[1]/input")
    grupo1.send_keys(grupo.value)

    time.sleep(1)

    llevar1 = driver.find_element_by_xpath("//*[@id='i33']/div[2]")
    llevar1.click()

    enviar1 = driver.find_element_by_xpath("//*[@id='mG61Hd']/div[2]/div/div[3]/div[1]/div/div/span/span")
    enviar1.click()
    
    
    time.sleep(3)
    

